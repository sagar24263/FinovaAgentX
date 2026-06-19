import json
import os
import tempfile
from typing import Optional

from fastapi import APIRouter, File, Form, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import JSONResponse

from app.models.pdf_parser import PdfParserResponse
from app.services.pdf_parser_service import PdfParserService
from app.utils.logger import get_logger

logger = get_logger("pdf_parser_routes")

router = APIRouter()
pdf_parser_service = PdfParserService()

_PDF_MAGIC = b"%PDF"
_CHUNK_SIZE = 1024 * 1024


@router.post("/parse-pdf", response_model=PdfParserResponse)
async def parse_pdf(
    file: UploadFile = File(...),
    nfo_name: str = Form(..., description="Fund/NFO name (e.g. 'HDFC Life Top 500 Smart Value 50 Fund')"),
    insurer_name: str = Form(..., description="Insurer name (e.g. 'HDFC Life')"),
    insurer_id: Optional[int] = Form(None, description="Optional insurer ID"),
):
    if file.content_type != "application/pdf":
        return PdfParserResponse(isSuccess=False, errors=["Only PDF files are accepted"])

    tmp_path: str | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp_path = tmp.name
            first_chunk = True
            while True:
                chunk = await file.read(_CHUNK_SIZE)
                if not chunk:
                    break
                if first_chunk:
                    if not chunk.startswith(_PDF_MAGIC):
                        return PdfParserResponse(isSuccess=False, errors=["File is not a valid PDF"])
                    first_chunk = False
                tmp.write(chunk)

        source_pdf = os.path.basename(file.filename or "upload.pdf")
        logger.info(f"Processing PDF: {source_pdf}")

        entries = await run_in_threadpool(
            pdf_parser_service.process,
            tmp_path,
            source_pdf,
            nfo_name,
            insurer_name,
            insurer_id,
        )

        if not entries:
            return PdfParserResponse(isSuccess=False, errors=["No content extracted from PDF"])

        return PdfParserResponse(isSuccess=True, data=entries)

    except Exception as e:
        logger.error(f"PDF parsing failed: {e}", exc_info=True)
        return PdfParserResponse(isSuccess=False, errors=[str(e)])

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


@router.post("/parse-pdfs-bulk")
async def parse_pdfs_bulk(
    excel_file: UploadFile = File(..., description="Excel file with columns: file_name, nfo_name, insurer_name, insurer_id"),
    pdf_folder: str = Form(..., description="Folder path where PDFs are stored (e.g. C:/Users/.../One Pagers)"),
):
    """
    Bulk PDF parsing. Upload an Excel with metadata. PDFs are read from the specified folder.

    Excel columns:
    - file_name: PDF filename (must exist in pdf_folder)
    - nfo_name: fund/product name
    - insurer_name: insurer name
    - insurer_id: (optional) insurer ID

    Returns JSON with all chunks for all PDFs.
    """
    import openpyxl

    # Save and read Excel
    try:
        excel_bytes = await excel_file.read()
        excel_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_tmp.write(excel_bytes)
        excel_tmp.close()

        wb = openpyxl.load_workbook(excel_tmp.name, read_only=True)
        ws = wb.active

        # Parse header row
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        required = {"file_name", "nfo_name", "insurer_name"}
        if not required.issubset(set(headers)):
            return JSONResponse(
                status_code=400,
                content={"isSuccess": False, "errors": [f"Excel must have columns: file_name, nfo_name, insurer_name. Found: {headers}"]},
            )

        # Parse rows
        col_idx = {h: i for i, h in enumerate(headers)}
        rows_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            file_name = str(row[col_idx["file_name"]] or "").strip()
            if not file_name:
                continue
            rows_data.append({
                "file_name": file_name,
                "nfo_name": str(row[col_idx["nfo_name"]] or "").strip(),
                "insurer_name": str(row[col_idx["insurer_name"]] or "").strip(),
                "insurer_id": int(row[col_idx["insurer_id"]]) if "insurer_id" in col_idx and row[col_idx["insurer_id"]] else None,
            })

        wb.close()
        os.remove(excel_tmp.name)

    except Exception as e:
        logger.error(f"Failed to parse Excel: {e}", exc_info=True)
        return JSONResponse(
            status_code=400,
            content={"isSuccess": False, "errors": [f"Failed to parse Excel: {e}"]},
        )

    if not rows_data:
        return JSONResponse(status_code=400, content={"isSuccess": False, "errors": ["Excel has no data rows"]})

    # Validate folder
    folder = pdf_folder.strip()
    if not os.path.isdir(folder):
        return JSONResponse(
            status_code=400,
            content={"isSuccess": False, "errors": [f"Folder not found: {folder}"]},
        )

    # Process each PDF
    all_results = []
    errors = []

    for row_meta in rows_data:
        filename = row_meta["file_name"]
        pdf_path = os.path.join(folder, filename)

        if not os.path.isfile(pdf_path):
            errors.append(f"File not found: {filename}")
            continue

        try:
            entries = await run_in_threadpool(
                pdf_parser_service.process,
                pdf_path,
                filename,
                row_meta["nfo_name"],
                row_meta["insurer_name"],
                row_meta["insurer_id"],
            )

            if entries:
                all_results.extend(entries)
                logger.info(f"Processed '{filename}': {len(entries)} chunks")
            else:
                errors.append(f"No content extracted from '{filename}'")

        except Exception as e:
            logger.error(f"Failed to process '{filename}': {e}", exc_info=True)
            errors.append(f"Error processing '{filename}': {e}")

    # Write results to JSON file in the same folder
    output_file = os.path.join(folder, "parsed_chunks_output.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    logger.info(f"Results written to: {output_file}")

    return JSONResponse(content={
        "isSuccess": len(all_results) > 0,
        "total_pdfs_processed": len(rows_data),
        "total_chunks": len(all_results),
        "output_file": output_file,
        "errors": errors if errors else None,
        "data": all_results,
    })
